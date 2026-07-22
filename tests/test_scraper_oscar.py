"""Tests for the scraper module — OSCAR side only (eoPortal detail is
flaky due to Cloudflare and is covered by an integration smoke test)."""

import os
import sys

import pytest

# Only run these if the optional packages are available
openpyxl = pytest.importorskip("openpyxl")
requests = pytest.importorskip("requests")

from core import scraper  # noqa: E402


def test_parse_oscar_xlsx_fixture(tmp_path):
    """Round-trip a minimal xlsx fixture through the parser."""
    import openpyxl as _opx
    wb = _opx.Workbook()
    ws = wb.active
    ws.append([
        "Acronym", "SAT ID", "Launch", "(expected) EOL",
        "Satellite Programme", "Agencies", "Orbit", "Altitude",
        "Longitude", "Inclination", "Ect", "Sat status", "Payload",
        "WIGOS Station Identifier", "Link", "Last update",
    ])
    # Note: openpyxl writes literal "\n" (one cell with a newline) only if
    # we pass it as a real \n in the Python string. The "\n" inside the
    # Python string literal is a real newline.
    ws.append([
        "TEST-SAT-1", 99999, "01 Jan 2020", "≥2025",
        "Test Programme", "NASA\nJAXA", "SunSync", "500 km",
        None, "98°", "10:30 desc", "Operational",
        "MSI (TEST-SAT-1)\nSAR",
        "0-20000-0-", None, "2025-01-01 00:00:00",
    ])
    p = tmp_path / "fixture.xlsx"
    wb.save(p)
    # Verify what openpyxl wrote
    check = _opx.load_workbook(p, read_only=True, data_only=True)
    chk_ws = check.active
    for row in chk_ws.iter_rows(values_only=True):
        if row[0] and "NASA" in str(row[0]):
            print(f"DEBUG: agencies cell = {row[5]!r}")
            break
    with open(p, "rb") as f:
        recs = scraper.parse_oscar_xlsx(f.read())
    assert len(recs) == 1
    r = recs[0]
    assert r["acronym"] == "TEST-SAT-1"
    assert r["sat_id"] == 99999
    assert r["agencies"] == ["NASA", "JAXA"]
    assert r["instruments"] == ["MSI", "SAR"]
    assert r["orbit"] == "SunSync"
    assert r["altitude"] == "500 km"
    assert r["status"] == "Operational"
    assert r["detail_url"] == "https://space.oscar.wmo.int/satellites/view/99999"


def test_normalize_text_handles_nbsp():
    assert scraper._normalize_text(" 25\xa0Aug\xa01997 ") == "25 Aug 1997"
    assert scraper._normalize_text("") is None
    assert scraper._normalize_text(None) is None
